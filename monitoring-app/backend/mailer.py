"""
mailer.py — Report generation and SMTP delivery for HyperMonitor.

Two report artefacts are built and attached to every email:

1. dashboard_summary.csv
   One row per hypervisor host with live CPU %, RAM used/total, storage
   used/total, VM count, and status.

2. vm_inventory_<server_name>.csv  (one file per host)
   Full VM inventory for that host: VM name, IP, power state, CPU %,
   vCPUs, RAM used / total, owner, creation date, purpose.

The HTML email body contains the same dashboard summary as an inline table
so recipients get an at-a-glance view without opening any attachment.

SMTP modes
──────────
  "smtps"    — SSL/TLS from the start (port 465).
  "starttls" — Plain connect then STARTTLS upgrade (port 587).
  "plain"    — No encryption (port 25 internal/corporate relay). Optional auth.

The smtp_mode field takes precedence. The legacy use_tls bool is still
honoured: use_tls=True → "smtps", use_tls=False with no smtp_mode → "starttls".
"""

from __future__ import annotations

import io
import logging
import re
import smtplib
import ssl
from collections import defaultdict
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# Shared openpyxl style helpers
# ──────────────────────────────────────────────────────────────────────────────

# Colour palette
_C_HEADER_BG   = "0F172A"   # dark navy  — sheet title row
_C_SECTION_BG  = "1E40AF"   # royal blue — section heading rows
_C_COL_HDR_BG  = "DBEAFE"   # light blue — column header rows
_C_COL_HDR_FG  = "1E3A5F"   # dark blue  — column header text
_C_ALT_ROW     = "F0F7FF"   # faint blue — alternating data rows
_C_CRITICAL    = "DC2626"   # red
_C_WARNING     = "D97706"   # amber
_C_ONLINE      = "16A34A"   # green
_C_STOPPED     = "6B7280"   # grey

_THIN = Side(style="thin", color="CBD5E1")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def _hdr_font(size: int = 11, bold: bool = True, color: str = "FFFFFF") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _body_font(size: int = 10, bold: bool = False, color: str = "1F2328") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _write_title_row(ws, text: str, ncols: int, row: int) -> None:
    """Merge across all columns and write a big title row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = _hdr_font(size=13, color="FFFFFF")
    cell.fill      = _fill(_C_HEADER_BG)
    cell.alignment = _left()
    cell.border    = _CELL_BORDER
    ws.row_dimensions[row].height = 22


def _write_section_heading(ws, text: str, ncols: int, row: int) -> None:
    """Merge across all columns and write a blue section heading."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = _hdr_font(size=11, color="FFFFFF")
    cell.fill      = _fill(_C_SECTION_BG)
    cell.alignment = _left()
    cell.border    = _CELL_BORDER
    ws.row_dimensions[row].height = 18


def _write_col_headers(ws, headers: List[str], row: int) -> None:
    """Write styled column header row."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = Font(name="Calibri", size=10, bold=True, color=_C_COL_HDR_FG)
        cell.fill      = _fill(_C_COL_HDR_BG)
        cell.alignment = _center()
        cell.border    = _CELL_BORDER
    ws.row_dimensions[row].height = 16


def _write_data_row(ws, values: List, row: int, alt: bool = False,
                    status_col: int | None = None) -> None:
    """Write a styled data row; optionally colour the status cell."""
    bg = _C_ALT_ROW if alt else "FFFFFF"
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = _body_font()
        cell.fill      = _fill(bg)
        cell.alignment = _left()
        cell.border    = _CELL_BORDER
        # Colour the status column text
        if status_col and col == status_col and isinstance(val, str):
            v = val.lower()
            if v in ("critical",):
                cell.font = _body_font(bold=True, color=_C_CRITICAL)
            elif v in ("warning",):
                cell.font = _body_font(bold=True, color=_C_WARNING)
            elif v in ("online", "running"):
                cell.font = _body_font(bold=True, color=_C_ONLINE)
            elif v in ("stopped", "offline"):
                cell.font = _body_font(bold=True, color=_C_STOPPED)


def _autofit(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Approximate column auto-fit based on cell content length."""
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


# ──────────────────────────────────────────────────────────────────────────────
# Excel report builder  (one .xlsx with multiple sheets)
# ──────────────────────────────────────────────────────────────────────────────

def _build_xlsx(servers: List[Dict], vms: List[Dict]) -> bytes:
    """
    Build and return a fully-styled Excel workbook (.xlsx) as bytes.

    Sheet layout:
      • "Dashboard Summary"   — KPI block + host utilisation table
      • One sheet per server  — VM inventory for that host
    """
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb  = Workbook()

    # ── Sheet 1: Dashboard Summary ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard Summary"
    ws.sheet_view.showGridLines = False

    total_vms   = sum(s.get("vm_count", 0) for s in servers)
    running_vms = sum(1 for v in vms if v.get("power_state") == "running")
    critical_ct = sum(1 for s in servers if s.get("status") == "critical")
    warning_ct  = sum(1 for s in servers if s.get("status") == "warning")
    avg_cpu     = (
        round(sum(s.get("cpu_usage_pct", 0) for s in servers) / len(servers), 1)
        if servers else 0
    )

    NCOLS_DASH = 13   # number of data columns in the host table
    r = 1

    # Report title
    _write_title_row(ws, "HyperMonitor — Infrastructure Dashboard Report", NCOLS_DASH, r)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS_DASH)
    gen_cell = ws.cell(row=r, column=1, value=f"Generated: {now}")
    gen_cell.font      = _body_font(size=10, color="57606A")
    gen_cell.alignment = _left()
    r += 2

    # ── Executive Summary section ─────────────────────────────────────────────
    _write_section_heading(ws, "Executive Summary", NCOLS_DASH, r); r += 1

    kpi_labels = [
        ("Total Hypervisor Hosts",  len(servers)),
        ("Total Virtual Machines",  total_vms),
        ("Running VMs",             running_vms),
        ("Average CPU Utilisation", f"{avg_cpu}%"),
        ("Hosts in Critical State", critical_ct),
        ("Hosts in Warning State",  warning_ct),
        ("Hosts Online",            len(servers) - critical_ct - warning_ct),
    ]
    for label, val in kpi_labels:
        lbl_cell = ws.cell(row=r, column=1, value=label)
        lbl_cell.font      = _body_font(bold=True)
        lbl_cell.fill      = _fill("F8FAFC")
        lbl_cell.alignment = _left()
        lbl_cell.border    = _CELL_BORDER

        val_cell = ws.cell(row=r, column=2, value=val)
        val_cell.font      = _body_font(bold=True, color="1E40AF")
        val_cell.fill      = _fill("F8FAFC")
        val_cell.alignment = _center()
        val_cell.border    = _CELL_BORDER

        # Merge columns 3→NCOLS for the label row so it looks clean
        ws.merge_cells(start_row=r, start_column=3,
                       end_row=r, end_column=NCOLS_DASH)
        ws.cell(row=r, column=3).fill   = _fill("F8FAFC")
        ws.cell(row=r, column=3).border = _CELL_BORDER
        r += 1

    r += 1  # blank spacer

    # ── Host Utilisation Detail section ──────────────────────────────────────
    _write_section_heading(ws, "Host Utilisation Detail", NCOLS_DASH, r); r += 1

    host_headers = [
        "Host Name", "IP Address", "Hypervisor Type",
        "CPU %", "CPU Cores",
        "RAM Used (GB)", "RAM Total (GB)", "RAM %",
        "Storage Used (TB)", "Storage Total (TB)", "Storage %",
        "VM Count", "Status",
    ]
    _write_col_headers(ws, host_headers, r); r += 1

    for i, s in enumerate(servers):
        _write_data_row(
            ws,
            [
                s.get("display_name", ""),
                s.get("ip_address", ""),
                s.get("hypervisor_type", ""),
                s.get("cpu_usage_pct", 0),
                s.get("cpu_cores", 0),
                s.get("ram_used_gb", 0),
                s.get("ram_total_gb", 0),
                s.get("ram_usage_pct", 0),
                s.get("storage_used_tb", 0),
                s.get("storage_total_tb", 0),
                s.get("storage_usage_pct", 0),
                s.get("vm_count", 0),
                s.get("status", "").upper(),
            ],
            row=r,
            alt=(i % 2 == 1),
            status_col=13,   # "Status" is column 13
        )
        r += 1

    # Fix column widths
    ws.column_dimensions["A"].width = 20   # Host Name
    ws.column_dimensions["B"].width = 16   # IP Address
    ws.column_dimensions["C"].width = 18   # Hypervisor Type
    for col in ["D","E","F","G","H","I","J","K","L"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["M"].width = 12   # Status

    # Freeze panes below column headers so scrolling keeps headers visible
    ws.freeze_panes = "A5"

    # ── Sheet per server: VM Inventory ────────────────────────────────────────
    server_map: Dict[str, str] = {s["server_id"]: s["display_name"] for s in servers}
    vms_by_server: Dict[str, List[Dict]] = defaultdict(list)
    for v in vms:
        vms_by_server[v.get("host_server_id", "unknown")].append(v)

    NCOLS_VM = 13

    for sid, vm_list in vms_by_server.items():
        sname_full  = server_map.get(sid, sid)
        # Sheet names max 31 chars, no special chars
        sheet_title = re.sub(r"[\\/*?:\[\]]", "_", sname_full)[:31]

        ws2 = wb.create_sheet(title=sheet_title)
        ws2.sheet_view.showGridLines = False

        running_ct = sum(1 for v in vm_list if v.get("power_state") == "running")
        stopped_ct = sum(1 for v in vm_list if v.get("power_state") == "stopped")
        avg_cpu_vm = (
            round(sum(v.get("cpu_usage_pct", 0) for v in vm_list) / len(vm_list), 1)
            if vm_list else 0
        )

        r2 = 1
        _write_title_row(ws2,
                         f"HyperMonitor — VM Inventory: {sname_full}",
                         NCOLS_VM, r2); r2 += 1

        ws2.merge_cells(start_row=r2, start_column=1,
                        end_row=r2, end_column=NCOLS_VM)
        gc = ws2.cell(row=r2, column=1, value=f"Generated: {now}")
        gc.font = _body_font(size=10, color="57606A"); gc.alignment = _left()
        r2 += 2

        # VM Summary block
        _write_section_heading(ws2, "VM Summary", NCOLS_VM, r2); r2 += 1
        vm_kpis = [
            ("Total VMs",        len(vm_list)),
            ("Running",          running_ct),
            ("Stopped",          stopped_ct),
            ("Average CPU %",    f"{avg_cpu_vm}%"),
        ]
        for label, val in vm_kpis:
            lc = ws2.cell(row=r2, column=1, value=label)
            lc.font = _body_font(bold=True); lc.fill = _fill("F8FAFC")
            lc.alignment = _left(); lc.border = _CELL_BORDER

            vc = ws2.cell(row=r2, column=2, value=val)
            vc.font = _body_font(bold=True, color="1E40AF"); vc.fill = _fill("F8FAFC")
            vc.alignment = _center(); vc.border = _CELL_BORDER

            ws2.merge_cells(start_row=r2, start_column=3,
                            end_row=r2, end_column=NCOLS_VM)
            ws2.cell(row=r2, column=3).fill   = _fill("F8FAFC")
            ws2.cell(row=r2, column=3).border = _CELL_BORDER
            r2 += 1

        r2 += 1

        # VM Inventory table
        _write_section_heading(ws2, "VM Inventory Detail", NCOLS_VM, r2); r2 += 1
        vm_headers = [
            "VM Name", "IP Address", "Hypervisor Type", "Power State",
            "CPU %", "vCPUs",
            "RAM Used (GB)", "RAM Total (GB)", "RAM %",
            "Owner", "Creation Date", "Purpose", "Status",
        ]
        _write_col_headers(ws2, vm_headers, r2); r2 += 1

        for i, v in enumerate(vm_list):
            _write_data_row(
                ws2,
                [
                    v.get("vm_name", ""),
                    v.get("ip_address", ""),
                    v.get("hypervisor_type", ""),
                    v.get("power_state", "").upper(),
                    v.get("cpu_usage_pct", 0),
                    v.get("cpu_cores", 0),
                    v.get("ram_used_gb", 0),
                    v.get("ram_total_gb", 0),
                    v.get("ram_usage_pct", 0),
                    v.get("owner_name", ""),
                    v.get("creation_date", ""),
                    v.get("purpose", ""),
                    v.get("status", "").upper(),
                ],
                row=r2,
                alt=(i % 2 == 1),
                status_col=13,  # "Status" is column 13
            )
            r2 += 1

        # Column widths for VM sheet
        ws2.column_dimensions["A"].width = 24  # VM Name
        ws2.column_dimensions["B"].width = 16  # IP Address
        ws2.column_dimensions["C"].width = 18  # Hypervisor Type
        ws2.column_dimensions["D"].width = 12  # Power State
        for col in ["E","F","G","H","I"]:
            ws2.column_dimensions[col].width = 14
        ws2.column_dimensions["J"].width = 16  # Owner
        ws2.column_dimensions["K"].width = 14  # Creation Date
        ws2.column_dimensions["L"].width = 20  # Purpose
        ws2.column_dimensions["M"].width = 12  # Status

        ws2.freeze_panes = "A5"

    # Serialise to bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# HTML report builder  (used by both email and the download endpoint)
# ──────────────────────────────────────────────────────────────────────────────

def _status_colour(status: str) -> str:
    return {"critical": "#dc2626", "warning": "#d97706", "online": "#16a34a"}.get(
        status.lower(), "#6b7280"
    )


def _build_html_body(servers: List[Dict], vms: List[Dict], generated_at: str) -> str:
    total_hosts = len(servers)
    total_vms   = len(vms)
    running_vms = sum(1 for v in vms if v.get("power_state") == "running")
    critical    = sum(1 for s in servers if s.get("status") == "critical")
    warning     = sum(1 for s in servers if s.get("status") == "warning")
    avg_cpu     = (
        round(sum(s.get("cpu_usage_pct", 0) for s in servers) / len(servers), 1)
        if servers else 0
    )

    rows_html = ""
    for s in servers:
        colour = _status_colour(s.get("status", ""))
        rows_html += f"""
        <tr>
          <td>{s.get('display_name','')}</td>
          <td style="font-family:monospace">{s.get('ip_address','')}</td>
          <td>{s.get('hypervisor_type','')}</td>
          <td>{s.get('cpu_usage_pct',0)}%</td>
          <td>{s.get('ram_used_gb',0)} / {s.get('ram_total_gb',0)} GB</td>
          <td>{s.get('storage_used_tb',0)} / {s.get('storage_total_tb',0)} TB</td>
          <td>{s.get('vm_count',0)}</td>
          <td><span style="color:{colour};font-weight:600">{s.get('status','').upper()}</span></td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"/>
<style>
  body {{ font-family: -apple-system, 'Segoe UI', sans-serif; font-size: 14px;
          color: #1f2328; margin: 0; padding: 0; background: #f7f8fa; }}
  .wrap {{ max-width: 760px; margin: 24px auto; background: #fff;
           border: 1px solid #e5e7eb; border-radius: 10px; overflow: hidden; }}
  .header {{ background: #0f172a; color: #fff; padding: 20px 28px; }}
  .header h1 {{ margin: 0; font-size: 18px; font-weight: 700; }}
  .header p  {{ margin: 4px 0 0; font-size: 12px; color: #94a3b8; }}
  .kpis {{ display: flex; gap: 0; border-bottom: 1px solid #e5e7eb; }}
  .kpi {{ flex: 1; padding: 16px 20px; border-right: 1px solid #e5e7eb; }}
  .kpi:last-child {{ border-right: none; }}
  .kpi-val {{ font-size: 26px; font-weight: 700; color: #1f2328; }}
  .kpi-lbl {{ font-size: 11px; color: #57606a; text-transform: uppercase;
               letter-spacing: .05em; }}
  .section {{ padding: 20px 28px; }}
  .section h2 {{ font-size: 14px; font-weight: 700; margin: 0 0 12px;
                  color: #1f2328; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #f7f8fa; padding: 8px 10px; text-align: left;
        font-size: 11px; font-weight: 700; text-transform: uppercase;
        letter-spacing: .04em; color: #57606a;
        border-bottom: 1px solid #e5e7eb; }}
  td {{ padding: 9px 10px; border-bottom: 1px solid #f0f0f0; }}
  tr:last-child td {{ border-bottom: none; }}
  .footer {{ padding: 14px 28px; background: #f7f8fa;
             border-top: 1px solid #e5e7eb; font-size: 11px; color: #94a3b8; }}
  .alert-box {{ margin: 0 28px 16px; padding: 12px 16px; border-radius: 8px;
                background: #fef2f2; border: 1px solid #fecaca;
                color: #b91c1c; font-size: 13px; }}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <h1>HyperMonitor — Infrastructure Report</h1>
    <p>Generated: {generated_at} &nbsp;·&nbsp; {total_hosts} host{'s' if total_hosts != 1 else ''} monitored</p>
  </div>

  <div class="kpis">
    <div class="kpi"><div class="kpi-val">{total_hosts}</div><div class="kpi-lbl">Total Hosts</div></div>
    <div class="kpi"><div class="kpi-val">{total_vms}</div><div class="kpi-lbl">Total VMs</div></div>
    <div class="kpi"><div class="kpi-val">{running_vms}</div><div class="kpi-lbl">Running VMs</div></div>
    <div class="kpi"><div class="kpi-val">{avg_cpu}%</div><div class="kpi-lbl">Avg CPU</div></div>
    <div class="kpi"><div class="kpi-val" style="color:{'#dc2626' if critical else '#16a34a'}">{critical}</div><div class="kpi-lbl">Critical</div></div>
    <div class="kpi"><div class="kpi-val" style="color:{'#d97706' if warning else '#16a34a'}">{warning}</div><div class="kpi-lbl">Warning</div></div>
  </div>

  {'<div class="alert-box">&#9888; ' + str(critical) + ' host(s) in CRITICAL state — immediate action required.</div>' if critical else ''}

  <div class="section">
    <h2>Dashboard Summary</h2>
    <table>
      <thead><tr>
        <th>Host</th><th>IP</th><th>Hypervisor</th>
        <th>CPU %</th><th>RAM</th><th>Storage</th><th>VMs</th><th>Status</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>

  <div class="footer">
    A fully-formatted Excel workbook (hypermonitor_report.xlsx) is attached.<br>
    This report was sent automatically by HyperMonitor.
  </div>
</div>
</body></html>"""


# ──────────────────────────────────────────────────────────────────────────────
# Public helper — build the HTML report as raw bytes (used by download endpoint)
# ──────────────────────────────────────────────────────────────────────────────

def build_html_report(servers: List[Dict], vms: List[Dict]) -> bytes:
    """
    Return the full HTML report as UTF-8 bytes.
    Used by GET /api/reports/report.html for direct browser download.
    """
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return _build_html_body(servers, vms, generated_at).encode("utf-8")


# ──────────────────────────────────────────────────────────────────────────────
# SMTP send
# ──────────────────────────────────────────────────────────────────────────────

def send_report(
    smtp_host: str,
    smtp_port: int,
    smtp_user: str,
    smtp_password: str,
    use_tls: bool,
    from_address: str,
    recipients: List[str],
    servers: List[Dict],
    vms: List[Dict],
    smtp_mode: Optional[str] = None,    # "smtps" | "starttls" | "plain"
    report_format: str = "both",        # "html" | "csv" | "both"
) -> None:
    """
    Build the report in the requested format and deliver it via SMTP.

    report_format values:
      "html"  — email body is the rich HTML report; no CSV attachments
      "csv"   — plain-text email with only CSV file attachments; no HTML body
      "both"  — HTML body + all CSV attachments (default / original behaviour)

    smtp_mode values:
      "smtps"    — SSL/TLS from the start (port 465)
      "starttls" — Plain TCP connect then STARTTLS upgrade (port 587)
      "plain"    — No encryption; suitable for internal port-25 relays
    """
    if not recipients:
        raise ValueError("No recipients configured.")
    if not smtp_host:
        raise ValueError("SMTP host is not configured.")

    # Normalise — default to "both" for any unrecognised value
    fmt = report_format if report_format in ("html", "csv", "both") else "both"

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    subject = f"HyperMonitor Report — {generated_at}"

    # ── Build email ──────────────────────────────────────────────────────────
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = from_address or smtp_user
    msg["To"]      = ", ".join(recipients)

    # ── Attach HTML body (html or both) ──────────────────────────────────────
    if fmt in ("html", "both"):
        html_body = _build_html_body(servers, vms, generated_at)
        msg.attach(MIMEText(html_body, "html", "utf-8"))
    else:
        # xlsx-only — add a minimal plain-text body so the email isn't empty
        plain = (
            f"HyperMonitor Infrastructure Report\n"
            f"Generated: {generated_at}\n\n"
            f"Hosts monitored: {len(servers)}\n"
            f"Total VMs: {len(vms)}\n\n"
            f"Full data is in the attached Excel report."
        )
        msg.attach(MIMEText(plain, "plain", "utf-8"))

    # ── Attach Excel workbook (csv or both) ──────────────────────────────────
    # A single .xlsx replaces the multiple CSV files — all sheets are included:
    #   Sheet 1: Dashboard Summary (KPI block + host utilisation table)
    #   Sheet N: VM inventory per host (one sheet per server)
    if fmt in ("csv", "both"):
        xlsx_bytes = _build_xlsx(servers, vms)
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(xlsx_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        "attachment", filename="hypermonitor_report.xlsx")
        msg.attach(part)

    # ── Send ─────────────────────────────────────────────────────────────────
    #
    # smtp_mode determines the connection strategy:
    #   "smtps"    — SSL wraps the entire connection from the start (port 465)
    #   "starttls" — Plain TCP connect then STARTTLS upgrade (port 587)
    #   "plain"    — No encryption at all (port 25 internal/corporate relay)
    #
    # If smtp_mode is not supplied, fall back to the legacy use_tls bool:
    #   use_tls=True  → "smtps"
    #   use_tls=False → "starttls"
    # ─────────────────────────────────────────────────────────────────────────
    mode = smtp_mode or ("smtps" if use_tls else "starttls")
    context = ssl.create_default_context()

    if mode == "smtps":
        with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context,
                              timeout=30) as srv:
            if smtp_user and smtp_password:
                srv.login(smtp_user, smtp_password)
            srv.sendmail(msg["From"], recipients, msg.as_string())

    elif mode == "starttls":
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as srv:
            srv.ehlo()
            srv.starttls(context=context)
            srv.ehlo()
            if smtp_user and smtp_password:
                srv.login(smtp_user, smtp_password)
            srv.sendmail(msg["From"], recipients, msg.as_string())

    else:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as srv:
            srv.ehlo()
            if smtp_user and smtp_password:
                srv.login(smtp_user, smtp_password)
            srv.sendmail(msg["From"], recipients, msg.as_string())

    log.info("Report sent to %s via %s:%s (mode=%s, format=%s)",
             recipients, smtp_host, smtp_port, mode, fmt)
